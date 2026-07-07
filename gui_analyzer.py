"""
Excel数据批量分析工具
功能：拖拽Excel文件 → 自动分析 → 生成统计结果Excel
"""
import os
import threading
from datetime import datetime
from collections import defaultdict

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DRAG_DROP = True
except ImportError:
    HAS_DRAG_DROP = False

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


PRIMARY_COLOR = "#2563EB"
PRIMARY_HOVER = "#1D4ED8"
SECONDARY_BG = "#F8FAFC"
BORDER_COLOR = "#E2E8F0"
TEXT_PRIMARY = "#1E293B"
TEXT_SECONDARY = "#64748B"
SUCCESS_COLOR = "#10B981"
ACCENT_BG = "#EFF6FF"


class ExcelAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 数据智能分析工具")
        self.root.geometry("680x600")
        self.root.resizable(False, False)
        self.root.configure(bg=SECONDARY_BG)

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status_text = tk.StringVar(value="就绪 — 将 Excel 文件拖入下方区域开始分析")
        self.is_processing = False
        self.drag_active = False

        self.setup_styles()
        self.build_ui()

        if HAS_DRAG_DROP:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.on_drop)

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')

        style.configure("Card.TFrame", background="white", relief="flat")
        style.configure("Title.TLabel", background=SECONDARY_BG,
                        foreground=TEXT_PRIMARY, font=("微软雅黑", 20, "bold"))
        style.configure("Subtitle.TLabel", background=SECONDARY_BG,
                        foreground=TEXT_SECONDARY, font=("微软雅黑", 10))

        style.configure("Section.TLabel", background="white",
                        foreground=TEXT_PRIMARY, font=("微软雅黑", 12, "bold"))
        style.configure("Field.TLabel", background="white",
                        foreground=TEXT_SECONDARY, font=("微软雅黑", 9))
        style.configure("Value.TLabel", background="white",
                        foreground=TEXT_PRIMARY, font=("微软雅黑", 10))

        style.configure("Primary.TButton",
                        font=("微软雅黑", 11, "bold"),
                        padding=(20, 10),
                        background=PRIMARY_COLOR,
                        foreground="white",
                        borderwidth=0,
                        focusthickness=0)
        style.map("Primary.TButton",
                  background=[("active", PRIMARY_HOVER), ("pressed", "#1E40AF")])

        style.configure("Ghost.TButton",
                        font=("微软雅黑", 10),
                        padding=(12, 6),
                        background=ACCENT_BG,
                        foreground=PRIMARY_COLOR,
                        borderwidth=0,
                        focusthickness=0)
        style.map("Ghost.TButton",
                  background=[("active", "#DBEAFE"), ("pressed", "#BFDBFE")])

        style.configure("TEntry",
                        fieldbackground="white",
                        bordercolor=BORDER_COLOR,
                        lightcolor=BORDER_COLOR,
                        darkcolor=BORDER_COLOR,
                        padding=6)

        style.configure("TProgressbar",
                        background=PRIMARY_COLOR,
                        troughcolor="#E2E8F0",
                        bordercolor="#E2E8F0",
                        lightcolor=PRIMARY_COLOR,
                        darkcolor=PRIMARY_COLOR,
                        thickness=8)

        style.configure("TFrame", background="white")
        style.configure("TLabelframe", background="white",
                        bordercolor=BORDER_COLOR, relief="solid")
        style.configure("TLabelframe.Label",
                        background="white",
                        foreground=TEXT_PRIMARY,
                        font=("微软雅黑", 10, "bold"))

    def build_ui(self):
        main = ttk.Frame(self.root, style="Card.TFrame")
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # ===== 标题区 =====
        header = ttk.Frame(main, style="Card.TFrame")
        header.pack(fill=tk.X, padx=24, pady=(20, 16))

        ttk.Label(header, text=" Excel 数据智能分析",
                  style="Title.TLabel").pack(side=tk.LEFT)
        ttk.Label(header, text="v1.0",
                  style="Subtitle.TLabel").pack(side=tk.LEFT, padx=(10, 0), pady=(12, 0))

        # ===== 拖拽区 =====
        self.drag_frame = tk.Frame(main, bg=ACCENT_BG,
                                   highlightthickness=2,
                                   highlightbackground="#93C5FD",
                                   highlightcolor=PRIMARY_COLOR,
                                   cursor="hand2")
        self.drag_frame.pack(fill=tk.X, padx=24, pady=(0, 16), ipady=30)

        self.drag_icon = tk.Label(self.drag_frame, text="📊",
                                  font=("Segoe UI Emoji", 36),
                                  bg=ACCENT_BG, fg=PRIMARY_COLOR)
        self.drag_icon.pack(pady=(20, 6))

        self.drag_title = tk.Label(self.drag_frame,
                                   text="将 Excel 文件拖拽到此处",
                                   font=("微软雅黑", 13, "bold"),
                                   bg=ACCENT_BG, fg=TEXT_PRIMARY)
        self.drag_title.pack()

        self.drag_hint = tk.Label(self.drag_frame,
                                  text="支持 .xlsx 格式 · 自动生成多维统计结果",
                                  font=("微软雅黑", 9),
                                  bg=ACCENT_BG, fg=TEXT_SECONDARY)
        self.drag_hint.pack(pady=(4, 20))

        self.drag_frame.bind("<Button-1>", lambda e: self.select_file())
        self.drag_icon.bind("<Button-1>", lambda e: self.select_file())
        self.drag_title.bind("<Button-1>", lambda e: self.select_file())
        self.drag_hint.bind("<Button-1>", lambda e: self.select_file())

        # ===== 文件信息卡片 =====
        info_card = ttk.Frame(main, style="Card.TFrame")
        info_card.pack(fill=tk.X, padx=24, pady=(0, 16))

        # 输入文件行
        row1 = ttk.Frame(info_card, style="Card.TFrame")
        row1.pack(fill=tk.X, padx=16, pady=(12, 6))
        ttk.Label(row1, text="📁  输入文件", style="Field.TLabel").pack(side=tk.LEFT)
        ttk.Label(row1, textvariable=self.input_path, style="Value.TLabel",
                  width=55, anchor="w").pack(side=tk.LEFT, padx=(8, 0))

        # 输出文件行
        row2 = ttk.Frame(info_card, style="Card.TFrame")
        row2.pack(fill=tk.X, padx=16, pady=(0, 12))
        ttk.Label(row2, text="📤  输出位置", style="Field.TLabel").pack(side=tk.LEFT)
        ttk.Label(row2, textvariable=self.output_path, style="Value.TLabel",
                  width=55, anchor="w").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(row2, text="更改", style="Ghost.TButton",
                   command=self.select_output).pack(side=tk.RIGHT)

        # ===== 操作按钮 =====
        btn_frame = ttk.Frame(main, style="Card.TFrame")
        btn_frame.pack(fill=tk.X, padx=24, pady=(0, 16))

        self.process_btn = ttk.Button(btn_frame, text="🚀  开始分析",
                                      style="Primary.TButton",
                                      command=self.start_processing)
        self.process_btn.pack(side=tk.LEFT)

        ttk.Button(btn_frame, text="选择文件", style="Ghost.TButton",
                   command=self.select_file).pack(side=tk.LEFT, padx=(10, 0))

        ttk.Button(btn_frame, text="打开输出目录", style="Ghost.TButton",
                   command=self.open_output_dir).pack(side=tk.RIGHT)

        # ===== 状态栏 =====
        status_card = ttk.Frame(main, style="Card.TFrame")
        status_card.pack(fill=tk.X, padx=24, pady=(0, 20))

        status_inner = ttk.Frame(status_card, style="Card.TFrame")
        status_inner.pack(fill=tk.X, padx=16, pady=10)

        self.status_icon = tk.Label(status_inner, text="●",
                                    font=("微软雅黑", 8),
                                    bg="white", fg=TEXT_SECONDARY)
        self.status_icon.pack(side=tk.LEFT)

        ttk.Label(status_inner, textvariable=self.status_text,
                  style="Value.TLabel").pack(side=tk.LEFT, padx=(8, 0))

        self.progress_bar = ttk.Progressbar(status_card, mode="indeterminate",
                                            length=600)
        self.progress_bar.pack(fill=tk.X, padx=16, pady=(0, 12))
        self.progress_bar.pack_forget()

    # ===== 拖拽与文件选择 =====
    def on_drop(self, event):
        if self.is_processing:
            messagebox.showwarning("提示", "正在处理中，请等待完成")
            return

        files = self.root.splitlist(event.data)
        if files:
            file_path = files[0]
            if file_path.lower().endswith('.xlsx'):
                self.load_file(file_path)
            else:
                messagebox.showwarning("提示", "请拖拽 .xlsx 格式的 Excel 文件")

    def select_file(self):
        if self.is_processing:
            return
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.load_file(file_path)

    def load_file(self, file_path):
        self.input_path.set(file_path)
        self.auto_set_output(file_path)
        file_name = os.path.basename(file_path)
        self.status_text.set(f"已加载文件：{file_name} — 点击开始分析")
        self.status_icon.configure(fg=SUCCESS_COLOR)

        self.drag_title.configure(text=f"✓  {file_name}")
        self.drag_hint.configure(text="文件已加载，点击下方按钮开始分析")
        self.drag_frame.configure(bg="#ECFDF5",
                                  highlightbackground="#6EE7B7",
                                  highlightcolor=SUCCESS_COLOR)
        self.drag_icon.configure(bg="#ECFDF5", fg=SUCCESS_COLOR)

    def auto_set_output(self, input_path):
        dir_name = os.path.dirname(input_path)
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_name = f"{base_name}_分析结果_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.output_path.set(os.path.join(dir_name, output_name))

    def select_output(self):
        file_path = filedialog.asksaveasfilename(
            title="选择输出文件位置",
            filetypes=[("Excel 文件", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if file_path:
            self.output_path.set(file_path)

    def open_output_dir(self):
        output = self.output_path.get().strip()
        if output:
            folder = os.path.dirname(output)
            if os.path.isdir(folder):
                os.startfile(folder)
            else:
                messagebox.showinfo("提示", "输出目录还不存在，分析完成后自动创建")
        else:
            messagebox.showinfo("提示", "请先选择或加载文件")

    # ===== 处理流程 =====
    def start_processing(self):
        if self.is_processing:
            return

        input_path = self.input_path.get().strip()
        output_path = self.output_path.get().strip()

        if not input_path:
            messagebox.showwarning("提示", "请先选择或拖拽 Excel 文件")
            return
        if not output_path:
            messagebox.showwarning("提示", "请设置输出文件位置")
            return
        if not os.path.exists(input_path):
            messagebox.showerror("错误", f"文件不存在：\n{input_path}")
            return

        self.is_processing = True
        self.process_btn.configure(state="disabled")
        self.progress_bar.pack(fill=tk.X, padx=16, pady=(0, 12))
        self.progress_bar.start()
        self.status_icon.configure(fg=PRIMARY_COLOR)

        thread = threading.Thread(target=self._do_process,
                                  args=(input_path, output_path))
        thread.daemon = True
        thread.start()

    def _do_process(self, input_path, output_path):
        try:
            self.status_text.set("步骤 1/3：读取 Excel 数据...")
            records = self.parse_excel(input_path)

            self.status_text.set(f"步骤 2/3：统计分析中（共 {len(records)} 条记录）...")
            summary = self.summarize_data(records)

            self.status_text.set("步骤 3/3：生成结果文件...")
            self.generate_output_excel(summary, records, output_path)

            file_name = os.path.basename(output_path)
            self.status_text.set(f"✓ 分析完成！共处理 {len(records)} 条记录 → {file_name}")
            self.status_icon.configure(fg=SUCCESS_COLOR)
            messagebox.showinfo("分析完成",
                                f"共处理 {len(records)} 条记录\n\n"
                                f"结果文件已保存至：\n{output_path}")
        except Exception as e:
            self.status_text.set(f"✗ 处理失败：{e}")
            self.status_icon.configure(fg="#EF4444")
            messagebox.showerror("处理失败", str(e))
        finally:
            self.is_processing = False
            self.process_btn.configure(state="normal")
            self.progress_bar.stop()
            self.progress_bar.pack_forget()

    # ===== 数据处理 =====
    def parse_excel(self, file_path):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError("Excel 数据为空（至少需要表头和一行数据）")
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = {}
            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                record[key] = val
            records.append(record)
        wb.close()
        return records

    def summarize_data(self, records):
        summary = {
            "total": len(records),
            "by_dept": defaultdict(list),
            "by_project": defaultdict(list),
            "by_status": defaultdict(int),
            "total_hours": 0.0,
        }
        for r in records:
            dept = r.get("部门", "未知")
            project = r.get("项目名称", "未知")
            status = r.get("完成状态", "未知")
            hours = r.get("工时(h)", 0) or 0
            summary["by_dept"][dept].append(r)
            summary["by_project"][project].append(r)
            summary["by_status"][status] += 1
            summary["total_hours"] += float(hours)
        summary["total_hours"] = round(summary["total_hours"], 1)
        return summary

    def generate_output_excel(self, summary, records, output_path):
        wb = Workbook()
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB",
                                  fill_type="solid")
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # 整体概况
        ws = wb.active
        ws.title = "整体概况"
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:C1')
        c = ws.cell(row=1, column=1, value="数据统计概况")
        c.font = Font(name="微软雅黑", bold=True, size=16, color="1E293B")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.cell(row=3, column=1, value="指标").fill = header_fill
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=2, value="数值").fill = header_fill
        ws.cell(row=3, column=2).font = header_font
        for col in (1, 2):
            ws.cell(row=3, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=col).border = thin_border

        metrics = [
            ("总记录数", summary["total"]),
            ("总工时 (小时)", summary["total_hours"]),
            ("部门数量", len(summary["by_dept"])),
            ("项目数量", len(summary["by_project"])),
        ]
        row = 4
        for k, v in metrics:
            ws.cell(row=row, column=1, value=k).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=2, value=v).font = Font(name="微软雅黑", size=10, bold=True)
            for col in (1, 2):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(vertical="center")
            if row % 2 == 0:
                for col in (1, 2):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            row += 1

        ws.cell(row=row, column=1, value="状态分布").font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        ws.cell(row=row, column=2, value="").fill = PatternFill(
            start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        for col in (1, 2):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        for status, cnt in sorted(summary["by_status"].items()):
            ws.cell(row=row, column=1, value=f"  {status}").font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=2, value=cnt).font = Font(name="微软雅黑", size=10)
            for col in (1, 2):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18

        # 部门统计
        self._write_stat_sheet(wb.create_sheet("部门统计"),
                               ["部门", "条目数", "总工时(h)", "完成率"],
                               summary["by_dept"], header_fill, header_font, thin_border)

        # 项目统计
        self._write_stat_sheet(wb.create_sheet("项目统计"),
                               ["项目名称", "条目数", "总工时(h)", "完成率"],
                               summary["by_project"], header_fill, header_font, thin_border, 24)

        # 待办事项
        ws_todo = wb.create_sheet("待办事项")
        ws_todo.sheet_view.showGridLines = False
        orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        todo_headers = ["项目", "工作内容", "负责人", "状态", "备注"]
        for col, h in enumerate(todo_headers, 1):
            c = ws_todo.cell(row=1, column=col, value=h)
            c.fill = orange_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        pending = [r for r in records if r.get("完成状态") in ("进行中", "已延期", "待开始")]
        row = 2
        for r in pending:
            values = [r.get("项目名称", ""), r.get("工作内容", ""),
                      r.get("员工", ""), r.get("完成状态", ""), r.get("备注", "")]
            for col, v in enumerate(values, 1):
                cell = ws_todo.cell(row=row, column=col, value=v)
                cell.font = Font(name="微软雅黑", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                for col in range(1, 6):
                    ws_todo.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
            row += 1
        for idx, w in enumerate([18, 24, 10, 10, 14], 1):
            ws_todo.column_dimensions[ws_todo.cell(row=1, column=idx).column_letter].width = w

        # 原始明细
        ws_detail = wb.create_sheet("原始明细")
        ws_detail.sheet_view.showGridLines = False
        if records:
            light_blue = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
            headers = list(records[0].keys())
            for col, h in enumerate(headers, 1):
                c = ws_detail.cell(row=1, column=col, value=h)
                c.fill = light_blue
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
            row = 2
            for r in records:
                for col, h in enumerate(headers, 1):
                    cell = ws_detail.cell(row=row, column=col, value=r.get(h, ""))
                    cell.font = Font(name="微软雅黑", size=10)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                if row % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        ws_detail.cell(row=row, column=col).fill = PatternFill(
                            start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
                row += 1
            for idx, w in enumerate([8, 14, 10, 10, 18, 20, 10, 10, 14], 1):
                if idx <= len(headers):
                    ws_detail.column_dimensions[ws_detail.cell(row=1, column=idx).column_letter].width = w

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)

    def _write_stat_sheet(self, ws, headers, data_dict, header_fill, header_font,
                          thin_border, name_width=16):
        ws.sheet_view.showGridLines = False
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        ws.row_dimensions[1].height = 28

        row = 2
        for key in sorted(data_dict.keys()):
            items = data_dict[key]
            cnt = len(items)
            hours = sum(float(r.get("工时(h)", 0) or 0) for r in items)
            done = sum(1 for r in items if r.get("完成状态") == "已完成")
            rate = f"{done / cnt * 100:.1f}%" if cnt else "0%"
            values = [key, cnt, round(hours, 1), rate]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            row += 1

        col_widths = [name_width, 10, 14, 10]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w


def main():
    try:
        if HAS_DRAG_DROP:
            root = TkinterDnD.Tk()
        else:
            root = tk.Tk()
        ExcelAnalyzerApp(root)
        root.mainloop()
    except Exception as e:
        print(f"启动失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
